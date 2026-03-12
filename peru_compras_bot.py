# -*- coding: utf-8 -*-
"""
peru_compras_bot.py
Automatiza la actualización de stock en catalogos.perucompras.gob.pe

Requisitos:
    pip install selenium pandas openpyxl

Uso:
    python peru_compras_bot.py
"""

import csv
import json
import logging
import os
import re
import sys
import threading
import time
import traceback
from datetime import datetime
from pathlib import Path
from queue import Empty, Queue

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

import pandas as pd
from openpyxl import Workbook as _OxlWorkbook
from openpyxl.styles import (
    PatternFill as _OxlFill, Font as _OxlFont,
    Alignment as _OxlAlignment, Border as _OxlBorder, Side as _OxlSide,
)
from selenium import webdriver
from selenium.common.exceptions import (
    NoAlertPresentException,
    NoSuchElementException,
    TimeoutException,
    UnexpectedAlertPresentException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

# ---------------------------------------------------------------------------
# CONFIGURACION - Ajusta estas variables segun tu entorno
# ---------------------------------------------------------------------------
EXCEL_PATH = Path(__file__).parent / "productos.xlsx"  # Ruta al archivo Excel
BASE_URL = "https://www.catalogos.perucompras.gob.pe"
LOGIN_URL = f"{BASE_URL}/AccesoGeneral"
MEJORA_URL = f"{BASE_URL}/MejoraBasica"

# Textos visibles de los selectores de filtro
ACUERDO_TEXTO = "EXT-CE-2022-5 COMPUTADORAS DE ESCRITORIO"
CATALOGO_TEXTO = "COMPUTADORAS DE ESCRITORIO"
CATEGORIA_TEXTO = "MONITOR"

# Tiempos de espera (segundos)
WAIT_NORMAL = 15
WAIT_LARGO = 30
WAIT_CORTO = 5
PAUSA_ENTRE_PRODUCTOS = 2  # Pausa entre iteraciones para no sobrecargar

# Estado para modo GUI
MODO_GUI = False
EVENTO_LOGIN = None
GUI_NOTIFICAR_LOGIN = None

# Control de ejecución (pausa / detenición / aprendizaje)
PAUSA_EVENTO = None       # threading.Event: set=corriendo, clear=en pausa
DETENER_EVENTO = None     # threading.Event: set=detener
ANALIZADOR = None         # instancia de AnalizadorFallos

# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(
            Path(__file__).parent / "bot.log", encoding="utf-8"
        ),
    ],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# REPORTE EXCEL
# ---------------------------------------------------------------------------
REPORTE_PATH = Path(__file__).parent / f"reporte_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
RESULTADOS: list = []


def nueva_ruta_reporte():
    return Path(__file__).parent / f"reporte_{datetime.now():%Y%m%d_%H%M%S}.xlsx"


def clasificar_error(mensaje: str) -> str:
    """Convierte un error técnico en una categoría comprensible para el usuario."""
    msg = str(mensaje).lower()
    if "no se encontraron resultados" in msg:
        return "Producto no encontrado en la tabla"
    if "n_stock" in msg or "campo de stock" in msg:
        return "Modal de stock no se abrió"
    if "timeout" in msg or "timed out" in msg:
        return "Tiempo de espera agotado"
    if "no such element" in msg:
        return "Elemento de página no encontrado"
    if "stale" in msg:
        return "Página cambió durante la operación"
    return "Error inesperado"


def registrar_resultado(parte: str, stock, estado: str, detalle: str = "", duracion: float = 0.0):
    """Registra el resultado de un producto en la lista interna."""
    RESULTADOS.append({
        "Parte": parte,
        "Stock": stock,
        "Estado": estado,
        "Tipo de Fallo": clasificar_error(detalle) if estado == "FALLO" else "",
        "Descripción": detalle,
        "Duración (seg)": round(duracion, 1),
    })
    simbolo = "OK" if estado == "EXITO" else "FALLO"
    log.info(f"  [{simbolo}] {parte} -> {detalle}")


def generar_plantilla_excel(destino: Path):
    """
    Crea un archivo Excel plantilla con instrucciones claras para que el usuario
    sepa cómo llenarlo correctamente antes de subirlo al bot.
    """
    wb = _OxlWorkbook()

    # ── helpers de estilo ────────────────────────────────────────────────
    def fill(color):
        return _OxlFill("solid", fgColor=color)

    def font(bold=False, size=10, color="000000", italic=False):
        return _OxlFont(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    borde = _OxlBorder(
        left=_OxlSide(style="thin", color="BFBFBF"),
        right=_OxlSide(style="thin", color="BFBFBF"),
        top=_OxlSide(style="thin", color="BFBFBF"),
        bottom=_OxlSide(style="thin", color="BFBFBF"),
    )
    ac = _OxlAlignment(horizontal="center", vertical="center", wrap_text=False)
    al = _OxlAlignment(horizontal="left", vertical="center", wrap_text=True)

    def c(ws, row, col, value, fll=None, fnt=None, aln=None):
        cell = ws.cell(row=row, column=col, value=value)
        if fll:
            cell.fill = fll
        if fnt:
            cell.font = fnt
        cell.border = borde
        cell.alignment = aln or al
        return cell

    # ── Hoja 1: Datos ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Productos"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    # Título
    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value = "Plantilla de Productos — Peru Compras Bot"
    cell.font = font(bold=True, size=13, color="FFFFFF")
    cell.fill = fill("00205B")
    cell.alignment = ac
    ws.row_dimensions[1].height = 24

    # Subtítulo / instrucción rápida
    ws.merge_cells("A2:B2")
    cell = ws["A2"]
    cell.value = "Completa las columnas 'Parte' y 'Stock' a partir de la fila 4. No borres los encabezados."
    cell.font = font(italic=True, size=9, color="595959")
    cell.alignment = _OxlAlignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 6

    # Encabezados
    for col, (txt, color_fll) in enumerate(
        [("Parte", "1F4E79"), ("Stock", "1F4E79")], start=1
    ):
        c(ws, 4, col, txt, fill(color_fll), font(bold=True, size=11, color="FFFFFF"), aln=ac)
    ws.row_dimensions[4].height = 20

    # Filas de ejemplo (se pueden borrar)
    ejemplos = [
        ("ABC-12345", 10),
        ("XYZ-67890", 0),
        ("MON-00001", 5),
    ]
    for i, (parte, stock) in enumerate(ejemplos, start=5):
        fll = fill("EBF3FB") if i % 2 == 0 else fill("FFFFFF")
        c(ws, i, 1, parte, fll, font(size=10, color="595959"))
        c(ws, i, 2, stock, fll, font(size=10, color="595959"), aln=ac)
        ws.row_dimensions[i].height = 16

    # ── Hoja 2: Instrucciones ────────────────────────────────────────────
    ws2 = wb.create_sheet("Instrucciones")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 60

    instrucciones = [
        ("INSTRUCCIONES PARA COMPLETAR EL ARCHIVO", "titulo"),
        ("", "sep"),
        ("1. Columna 'Parte'", "subtitulo"),
        ("   Escribe el número de parte exacto del producto tal como aparece en el portal.", "texto"),
        ("   Ejemplo: ABC-12345", "ejemplo"),
        ("", "sep"),
        ("2. Columna 'Stock'", "subtitulo"),
        ("   Escribe la cantidad de stock disponible. Solo números enteros (0, 1, 5, 100…).", "texto"),
        ("   Usa 0 para indicar sin stock. No uses comas ni puntos decimales.", "texto"),
        ("   Ejemplo: 10", "ejemplo"),
        ("", "sep"),
        ("3. Reglas importantes", "subtitulo"),
        ("   ✔  No borres ni renombres los encabezados 'Parte' y 'Stock'.", "texto"),
        ("   ✔  No dejes filas completamente vacías entre productos.", "texto"),
        ("   ✔  No uses hojas adicionales para los datos (usa siempre 'Productos').", "texto"),
        ("   ✔  Guarda el archivo en formato .xlsx antes de cargarlo en el bot.", "texto"),
        ("   ✗  No incluyas símbolos, espacios extra ni texto en la columna Stock.", "advertencia"),
        ("   ✗  No fusiones celdas ni apliques filtros en la hoja de datos.", "advertencia"),
        ("", "sep"),
        ("4. Cómo usar el archivo en el bot", "subtitulo"),
        ("   a) Guarda este archivo con el nombre que prefieras (ej: mis_productos.xlsx).", "texto"),
        ("   b) En la aplicación, haz clic en 'Seleccionar Excel' y elige este archivo.", "texto"),
        ("   c) Configura los filtros y haz clic en 'Iniciar automatización'.", "texto"),
        ("", "sep"),
        ("5. ¿Qué pasa si hay un error?", "subtitulo"),
        ("   El bot intentará procesar todos los productos.", "texto"),
        ("   Los que fallen quedan anotados en el reporte Excel con el motivo del error.", "texto"),
        ("   Puedes corregir esos productos y volver a ejecutar sólo con ellos.", "texto"),
    ]

    estilos = {
        "titulo":     (fill("00205B"), font(bold=True,  size=13, color="FFFFFF")),
        "sep":        (None,          None),
        "subtitulo":  (fill("D9E1F2"), font(bold=True,  size=10, color="1F4E79")),
        "texto":      (fill("FFFFFF"), font(size=10,    color="333333")),
        "ejemplo":    (fill("EBF3FB"), font(italic=True, size=10, color="2E75B6")),
        "advertencia":(fill("FFF2CC"), font(size=10,    color="9C6500")),
    }

    for i, (texto, estilo) in enumerate(instrucciones, start=1):
        fll, fnt = estilos.get(estilo, (None, None))
        ws2.merge_cells(f"A{i}:B{i}")
        cell = ws2[f"A{i}"]
        cell.value = texto
        if fll:
            cell.fill = fll
            cell.border = borde
        if fnt:
            cell.font = fnt
        cell.alignment = _OxlAlignment(horizontal="left", vertical="center", wrap_text=True)
        ws2.row_dimensions[i].height = 18 if texto else 8

    wb.save(destino)


def generar_reporte_excel(acuerdo_texto: str = "", catalogo_texto: str = "", categoria_texto: str = ""):
    """Genera un reporte Excel profesional con resumen, detalle por producto y gráficos."""
    from collections import Counter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.utils import get_column_letter
    Workbook = _OxlWorkbook
    PatternFill = _OxlFill
    Font = _OxlFont
    Alignment = _OxlAlignment
    Border = _OxlBorder
    Side = _OxlSide

    wb = Workbook()
    total = len(RESULTADOS)
    exitos = sum(1 for r in RESULTADOS if r["Estado"] == "EXITO")
    fallos = total - exitos

    # ── Helpers de estilo ────────────────────────────────────────────────
    def fill(color):
        return PatternFill("solid", fgColor=color)

    def font(bold=False, size=10, color="000000", italic=False):
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    borde = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    ac = Alignment(horizontal="center", vertical="center", wrap_text=False)
    al = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def celda(ws, row, col, value, fll=None, fnt=None, aln=None):  # noqa: E306
        c = ws.cell(row=row, column=col, value=value)
        if fll:
            c.fill = fll
        if fnt:
            c.font = fnt
        c.border = borde
        c.alignment = aln or ac
        return c

    # =====================================================================
    # HOJA 1: RESUMEN
    # =====================================================================
    ws_res = wb.active
    ws_res.title = "Resumen"
    ws_res.sheet_view.showGridLines = False
    ws_res.column_dimensions["A"].width = 30
    ws_res.column_dimensions["B"].width = 14
    ws_res.column_dimensions["C"].width = 14

    # Título
    ws_res.merge_cells("A1:C1")
    c = ws_res["A1"]
    c.value = "PERU COMPRAS BOT — Reporte de Actualización de Stock"
    c.font = font(bold=True, size=14, color="FFFFFF")
    c.fill = fill("00205B")
    c.alignment = ac
    ws_res.row_dimensions[1].height = 26

    # Fecha
    ws_res.merge_cells("A2:C2")
    c = ws_res["A2"]
    c.value = f"Generado el {datetime.now():%d/%m/%Y a las %H:%M:%S}"
    c.font = font(italic=True, size=9, color="595959")
    c.alignment = ac
    ws_res.row_dimensions[2].height = 14

    # Configuración
    ws_res.merge_cells("A3:C3")
    c = ws_res["A3"]
    c.value = f"Acuerdo: {acuerdo_texto}  |  Catálogo: {catalogo_texto}  |  Categoría: {categoria_texto}"
    c.font = font(italic=True, size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_res.row_dimensions[3].height = 14
    ws_res.row_dimensions[4].height = 8

    # Encabezado de tabla resumen
    for col, txt in enumerate(["Resultado", "Cantidad", "Porcentaje"], start=1):
        celda(ws_res, 5, col, txt, fill("1F4E79"), font(bold=True, size=11, color="FFFFFF"))
    ws_res.row_dimensions[5].height = 20

    pct_e = f"{exitos / total * 100:.1f}%" if total > 0 else "0.0%"
    pct_f = f"{fallos / total * 100:.1f}%" if total > 0 else "0.0%"

    for col, val in enumerate(["Exitosos ✔", exitos, pct_e], start=1):
        celda(ws_res, 6, col, val, fill("C6EFCE"), font(bold=True, size=10, color="375623"))
    ws_res.row_dimensions[6].height = 18

    for col, val in enumerate(["Fallidos ✘", fallos, pct_f], start=1):
        celda(ws_res, 7, col, val, fill("FFC7CE"), font(bold=True, size=10, color="9C0006"))
    ws_res.row_dimensions[7].height = 18

    for col, val in enumerate(["Total", total, "100%"], start=1):
        celda(ws_res, 8, col, val, fill("D9E1F2"), font(bold=True, size=10))
    ws_res.row_dimensions[8].height = 18

    # Gráfico de torta: Éxitos vs Fallos
    if total > 0:
        pie = PieChart()
        pie.title = "Resultados de Actualización"
        pie.style = 10
        pie.width = 13
        pie.height = 9
        data_ref = Reference(ws_res, min_col=2, min_row=5, max_row=7)
        cats_ref = Reference(ws_res, min_col=1, min_row=6, max_row=7)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        ws_res.add_chart(pie, "E1")

    # Desglose de tipos de fallo
    tipos = Counter(r["Tipo de Fallo"] for r in RESULTADOS if r["Estado"] == "FALLO")
    if tipos:
        ws_res.row_dimensions[10].height = 8
        for col, txt in enumerate(["Tipo de Fallo", "Cantidad"], start=1):
            celda(ws_res, 11, col, txt, fill("1F4E79"), font(bold=True, size=11, color="FFFFFF"))
        ws_res.row_dimensions[11].height = 20
        for i, (tipo, cnt) in enumerate(tipos.most_common(), start=12):
            celda(ws_res, i, 1, tipo, fill("FFC7CE"), font(size=10, color="9C0006"), aln=al)
            celda(ws_res, i, 2, cnt,  fill("FFC7CE"), font(bold=True, size=10, color="9C0006"))
            ws_res.row_dimensions[i].height = 16

        last_tipo_row = 11 + len(tipos)
        bar_tipos = BarChart()
        bar_tipos.type = "col"
        bar_tipos.title = "Detalle de Fallos por Categoría"
        bar_tipos.style = 10
        bar_tipos.y_axis.title = "Productos"
        bar_tipos.width = 13
        bar_tipos.height = 9
        data_b = Reference(ws_res, min_col=2, min_row=11, max_row=last_tipo_row)
        cats_b = Reference(ws_res, min_col=1, min_row=12, max_row=last_tipo_row)
        bar_tipos.add_data(data_b, titles_from_data=True)
        bar_tipos.set_categories(cats_b)
        ws_res.add_chart(bar_tipos, "E18")

    # =====================================================================
    # HOJA 2: DETALLE POR PRODUCTO
    # =====================================================================
    ws_det = wb.create_sheet("Detalle por Producto")
    ws_det.sheet_view.showGridLines = False

    ws_det.merge_cells("A1:G1")
    c = ws_det["A1"]
    c.value = "Detalle de Todos los Productos"
    c.font = font(bold=True, size=13, color="FFFFFF")
    c.fill = fill("1F4E79")
    c.alignment = ac
    ws_det.row_dimensions[1].height = 24

    det_headers = ["#", "Parte", "Stock", "Estado", "Tipo de Fallo", "Descripción del Error", "Duración (seg)"]
    for col, h in enumerate(det_headers, start=1):
        celda(ws_det, 2, col, h, fill("2E75B6"), font(bold=True, size=10, color="FFFFFF"))
    ws_det.row_dimensions[2].height = 18

    for i, r in enumerate(RESULTADOS, start=1):
        is_ok = r["Estado"] == "EXITO"
        fll = fill("C6EFCE") if is_ok else fill("FFC7CE")
        fnt_row = font(bold=True, size=9, color="375623" if is_ok else "9C0006")
        row = i + 2
        vals = [i, r["Parte"], r["Stock"], r["Estado"],
                r["Tipo de Fallo"] or "—", r["Descripción"], r["Duración (seg)"]]
        for col, val in enumerate(vals, start=1):
            celda(ws_det, row, col, val, fll, fnt_row, aln=al if col in (2, 5, 6) else ac)
        ws_det.row_dimensions[row].height = 16

    for col_letter, width in {"A": 5, "B": 22, "C": 8, "D": 10, "E": 30, "F": 42, "G": 14}.items():
        ws_det.column_dimensions[col_letter].width = width

    # Gráfico de barras: duración por producto
    if RESULTADOS:
        last_det = 2 + len(RESULTADOS)
        bar_dur = BarChart()
        bar_dur.type = "col"
        bar_dur.title = "Tiempo de Procesamiento por Producto (segundos)"
        bar_dur.style = 10
        bar_dur.y_axis.title = "Segundos"
        bar_dur.width = min(len(RESULTADOS) * 1.5 + 4, 28)
        bar_dur.height = 12
        data_dur = Reference(ws_det, min_col=7, min_row=2, max_row=last_det)
        cats_dur = Reference(ws_det, min_col=2, min_row=3, max_row=last_det)
        bar_dur.add_data(data_dur, titles_from_data=True)
        bar_dur.set_categories(cats_dur)
        ws_det.add_chart(bar_dur, f"A{last_det + 3}")

    # =====================================================================
    # HOJA 3: SOLO FALLIDOS
    # =====================================================================
    fallidos = [r for r in RESULTADOS if r["Estado"] == "FALLO"]
    ws_fail = wb.create_sheet("Solo Fallidos")
    ws_fail.sheet_view.showGridLines = False

    ws_fail.merge_cells("A1:F1")
    c = ws_fail["A1"]
    c.value = f"Productos Fallidos — {len(fallidos)} de {total}"
    c.font = font(bold=True, size=13, color="FFFFFF")
    c.fill = fill("9C0006")
    c.alignment = ac
    ws_fail.row_dimensions[1].height = 24

    if fallidos:
        fail_headers = ["#", "Parte", "Stock Intentado", "Tipo de Fallo", "Descripción del Error", "Duración (seg)"]
        for col, h in enumerate(fail_headers, start=1):
            celda(ws_fail, 2, col, h, fill("2E75B6"), font(bold=True, size=10, color="FFFFFF"))
        ws_fail.row_dimensions[2].height = 18

        for i, r in enumerate(fallidos, start=1):
            row = i + 2
            fll_alt = fill("FFE4E4") if i % 2 == 0 else fill("FFC7CE")
            vals = [i, r["Parte"], r["Stock"], r["Tipo de Fallo"], r["Descripción"], r["Duración (seg)"]]
            for col, val in enumerate(vals, start=1):
                celda(ws_fail, row, col, val, fll_alt, font(size=10, color="9C0006"),
                      aln=al if col in (2, 4, 5) else ac)
            ws_fail.row_dimensions[row].height = 18

        for col_letter, width in {"A": 5, "B": 22, "C": 14, "D": 32, "E": 48, "F": 14}.items():
            ws_fail.column_dimensions[col_letter].width = width
    else:
        ws_fail.merge_cells("A2:F2")
        c = ws_fail["A2"]
        c.value = "¡Sin fallos! Todos los productos se actualizaron correctamente."
        c.font = font(bold=True, size=11, color="375623")
        c.fill = fill("C6EFCE")
        c.alignment = ac
        ws_fail.row_dimensions[2].height = 22

    wb.save(REPORTE_PATH)
    log.info(f"Reporte Excel generado: {REPORTE_PATH}")
    return REPORTE_PATH


# ---------------------------------------------------------------------------
# HELPERS DE SELENIUM
# ---------------------------------------------------------------------------
def esperar_elemento(driver, by, valor, timeout=WAIT_NORMAL):
    """Espera a que un elemento sea visible y lo devuelve."""
    return WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((by, valor))
    )


def esperar_clickeable(driver, by, valor, timeout=WAIT_NORMAL):
    """Espera a que un elemento sea clickeable y lo devuelve."""
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, valor))
    )


def esperar_opciones_select(driver, select_id, timeout=WAIT_NORMAL):
    """Espera a que un <select> tenga mas de 1 opcion (la primera suele ser placeholder)."""
    def tiene_opciones(drv):
        sel = drv.find_element(By.ID, select_id)
        opciones = sel.find_elements(By.TAG_NAME, "option")
        return len(opciones) > 1

    WebDriverWait(driver, timeout).until(tiene_opciones)
    return Select(driver.find_element(By.ID, select_id))


def seleccionar_por_texto_parcial(select_obj: Select, texto_parcial: str):
    """Selecciona una opcion cuyo texto visible contenga el texto parcial dado."""
    for opcion in select_obj.options:
        if texto_parcial.upper() in opcion.text.upper():
            select_obj.select_by_visible_text(opcion.text)
            log.info(f"    Seleccionado: {opcion.text}")
            return True
    raise NoSuchElementException(
        f"No se encontro opcion que contenga '{texto_parcial}'"
    )


def aceptar_alerta(driver, timeout=WAIT_NORMAL):
    """Espera y acepta un alert/confirm del navegador."""
    try:
        WebDriverWait(driver, timeout).until(EC.alert_is_present())
        alerta = driver.switch_to.alert
        texto = alerta.text
        alerta.accept()
        log.info(f"    Alerta aceptada: {texto}")
        return texto
    except TimeoutException:
        log.warning("    No aparecio alerta en el tiempo esperado.")
        return None


def manejar_confirmacion_sweetalert(driver, timeout=WAIT_NORMAL):
    """
    Maneja dialogos tipo SweetAlert (muy comunes en este portal).
    Busca el boton de confirmacion "Si" / "Aceptar" / "Cerrar".
    """
    selectores_boton = [
        "//button[contains(@class,'swal2-confirm')]",
        "//button[contains(@class,'confirm')]",
        "//button[contains(text(),'Si')]",
        "//button[contains(text(),'Sí')]",
        "//button[contains(text(),'OK')]",
        "//button[contains(text(),'Aceptar')]",
        "//button[contains(text(),'Cerrar')]",
    ]
    for xpath in selectores_boton:
        try:
            btn = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.XPATH, xpath))
            )
            btn.click()
            log.info(f"    SweetAlert confirmado con: {xpath}")
            return True
        except TimeoutException:
            continue
    return False


def leer_opciones_select(driver, select_id, timeout=WAIT_NORMAL):
    """Lee las opciones de un <select> excluyendo el placeholder (value=0 o vacío)."""
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, select_id))
        )
        el = driver.find_element(By.ID, select_id)
        return [
            opt.text.strip()
            for opt in el.find_elements(By.TAG_NAME, "option")
            if opt.get_attribute("value") not in ("", "0") and opt.text.strip()
        ]
    except Exception:
        return []


# ---------------------------------------------------------------------------
# PASO 1: LOGIN Y VERIFICACION
# ---------------------------------------------------------------------------
def paso1_login(driver):
    """Carga la pagina de login y espera la verificacion de 2 pasos."""
    log.info("=" * 60)
    log.info("PASO 1: Login y verificacion")
    log.info("=" * 60)

    driver.get(LOGIN_URL)
    log.info(f"Pagina cargada: {LOGIN_URL}")

    # --- Esperar a que el usuario llene credenciales manualmente ---
    if MODO_GUI and EVENTO_LOGIN is not None:
        log.info("Esperando confirmación de login desde la interfaz...")
        if GUI_NOTIFICAR_LOGIN:
            GUI_NOTIFICAR_LOGIN()
        EVENTO_LOGIN.clear()
        EVENTO_LOGIN.wait()
    else:
        print("\n" + "=" * 60)
        print("ACCION MANUAL REQUERIDA")
        print("=" * 60)
        print("1. Ingresa tu RUC, usuario y contrasena en el navegador.")
        print("2. Haz clic en 'Ingresar'.")
        print("3. Si aparece la ventana de verificacion de codigo, solo presiona ENTER.")
        input(">>> Presiona ENTER aqui cuando hayas completado el login... ")

    # --- Saltar verificacion haciendo click atras ---
    log.info("Saltando verificacion de codigo con navegacion hacia atras...")
    driver.back()
    time.sleep(3)
    log.info(f"Navegacion hacia atras completada. URL actual: {driver.current_url}")

    # Esperar a que cargue el dashboard
    time.sleep(2)
    log.info(f"Login completado. URL actual: {driver.current_url}")


# ---------------------------------------------------------------------------
# PASO 2: NAVEGACION Y TRUCO DE RETROCESO
# ---------------------------------------------------------------------------
def paso2_navegacion(driver):
    """Navega a la seccion de Mejora Basica desbloqueando el menu."""
    log.info("=" * 60)
    log.info("PASO 2: Navegacion a Mejora de Precio y Existencias")
    log.info("=" * 60)

    # Truco de retroceso: volver atras para desbloquear el menu
    driver.back()
    time.sleep(2)
    driver.get(BASE_URL)
    time.sleep(2)
    log.info("Retroceso ejecutado para desbloquear menu.")

    # Intentar navegar por URL directa
    driver.get(MEJORA_URL)
    time.sleep(3)

    # Verificar que estamos en la pagina correcta
    if "MejoraBasica" in driver.current_url:
        log.info(f"Navegacion exitosa a: {driver.current_url}")
        return

    # Fallback: intentar por menu lateral
    log.info("Intentando navegacion por menu lateral...")
    try:
        menu_mejora = esperar_clickeable(
            driver,
            By.XPATH,
            "//a[contains(text(),'Mejora de ofertas')]"
            "|//span[contains(text(),'Mejora de ofertas')]"
            "|//li[contains(text(),'Mejora de ofertas')]",
        )
        menu_mejora.click()
        time.sleep(1)

        submenu = esperar_clickeable(
            driver,
            By.XPATH,
            "//a[contains(@href,'MejoraBasica')]"
            "|//a[contains(text(),'Precio y Existencias')]"
            "|//a[contains(text(),'Precio y existencias')]",
        )
        submenu.click()
        time.sleep(2)
        log.info(f"Navegacion por menu completada. URL: {driver.current_url}")
    except TimeoutException:
        log.warning(
            "No se pudo navegar por menu. Intentando URL directa nuevamente..."
        )
        driver.get(MEJORA_URL)
        time.sleep(3)


# ---------------------------------------------------------------------------
# PASO 3: CONFIGURACION DE FILTROS
# ---------------------------------------------------------------------------
def paso3_filtros(driver):
    """Selecciona Acuerdo Marco, Catalogo y Categoria."""
    log.info("=" * 60)
    log.info("PASO 3: Configuracion de filtros")
    log.info("=" * 60)

    # Acuerdo Marco
    log.info("  Seleccionando Acuerdo Marco...")
    select_acuerdo = esperar_opciones_select(driver, "ajaxAcuerdo", WAIT_LARGO)
    seleccionar_por_texto_parcial(select_acuerdo, ACUERDO_TEXTO)
    time.sleep(2)  # Esperar carga del siguiente select

    # Catalogo Electronico
    log.info("  Seleccionando Catalogo Electronico...")
    select_catalogo = esperar_opciones_select(driver, "ajaxCatalogo", WAIT_LARGO)
    seleccionar_por_texto_parcial(select_catalogo, CATALOGO_TEXTO)
    time.sleep(2)

    # Categoria
    log.info("  Seleccionando Categoria...")
    select_categoria = esperar_opciones_select(driver, "ajaxCategoria", WAIT_LARGO)
    seleccionar_por_texto_parcial(select_categoria, CATEGORIA_TEXTO)
    time.sleep(1)

    log.info("Filtros configurados correctamente.")


# ---------------------------------------------------------------------------
# PASO 4: BUCLE DE ACTUALIZACION DE STOCK
# ---------------------------------------------------------------------------
def paso4_actualizar_stock(driver, df: pd.DataFrame):
    """Itera sobre cada producto del DataFrame y actualiza el stock."""
    log.info("=" * 60)
    log.info(f"PASO 4: Actualizacion de stock ({len(df)} productos)")
    log.info("=" * 60)

    total = len(df)
    exitos = 0
    fallos = 0

    for idx, fila in df.iterrows():
        parte = str(fila["Parte"]).strip()
        stock = str(int(fila["Stock"]))
        log.info(f"\n--- Producto {idx + 1}/{total}: Parte={parte}, Stock={stock} ---")

        t_inicio = time.time()
        try:
            actualizar_producto(driver, parte, stock)
            registrar_resultado(parte, stock, "EXITO", "Stock actualizado correctamente",
                                duracion=time.time() - t_inicio)
            exitos += 1
        except Exception as e:
            log.error(f"  Error procesando {parte}: {e}")
            registrar_resultado(parte, stock, "FALLO", str(e),
                                duracion=time.time() - t_inicio)
            fallos += 1
            # Intentar recuperar el estado limpio para el siguiente producto
            try:
                recuperar_estado(driver)
            except Exception:
                log.warning("  No se pudo recuperar estado. Recargando pagina...")
                driver.get(MEJORA_URL)
                time.sleep(3)
                paso3_filtros(driver)

        time.sleep(PAUSA_ENTRE_PRODUCTOS)

    log.info("=" * 60)
    log.info(f"RESULTADO FINAL: {exitos} exitosos, {fallos} fallidos de {total} total")
    log.info(f"Reporte guardado en: {REPORTE_PATH}")
    log.info("=" * 60)


def actualizar_producto(driver, parte: str, stock: str):
    """Busca un producto por numero de parte y actualiza su stock."""
    extra_espera = ANALIZADOR.wait_extra() if ANALIZADOR else 0.0

    # 1. Escribir numero de parte en el campo de busqueda
    campo_busqueda = esperar_elemento(driver, By.ID, "C_Descripcion")
    campo_busqueda.clear()
    campo_busqueda.send_keys(parte)

    # 2. Click en buscar
    btn_buscar = esperar_clickeable(driver, By.ID, "btnBuscar")
    btn_buscar.click()
    log.info(f"  Busqueda lanzada para: {parte}")

    # 3. Esperar a que cargue la tabla de resultados
    time.sleep(3 + extra_espera)  # Pausa inicial para carga AJAX

    try:
        WebDriverWait(driver, WAIT_NORMAL).until(
            EC.presence_of_element_located((
                By.XPATH,
                "//table//tbody//tr[contains(@class,'row')]"
                "|//table//tbody//tr[td]"
                "|//div[contains(@class,'dataTables')]//tbody//tr"
            ))
        )
    except TimeoutException:
        raise TimeoutException(f"No se encontraron resultados para '{parte}'")

    # 4. Localizar el boton "Existencias" DENTRO DE LA TABLA de resultados
    # Ser muy específico: buscar en la tabla, el enlace con fnModificarStock
    xpath_existencias = (
        f"//table//tbody//tr//a[contains(@onclick,'fnModificarStock')]"
        f"|//table//tr//a[contains(@onclick,'fnModificarStock')]"
        f"|//div[contains(@class,'dataTables')]//a[contains(@onclick,'fnModificarStock')]"
    )

    btn_existencias = esperar_clickeable(
        driver, By.XPATH, xpath_existencias, WAIT_NORMAL
    )
    
    # Scroll al elemento antes de hacer click
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn_existencias)
    time.sleep(0.5)
    
    # DEBUGGING: Capturar info del botón encontrado
    try:
        tag_name = btn_existencias.tag_name
        onclick_value = btn_existencias.get_attribute("onclick")
        text = btn_existencias.text
        html = btn_existencias.get_attribute("outerHTML")
        
        log.info(f"  Botón encontrado (VERIFICAR):")
        log.info(f"    Tag: {tag_name}")
        log.info(f"    Texto: {text}")
        log.info(f"    onclick: {onclick_value[:100] if onclick_value else 'None'}")
        log.info(f"    HTML: {html[:200] if len(html) > 200 else html}")
    except Exception as e:
        log.error(f"  Error obteniendo info del botón: {e}")
    
    # NUEVA ESTRATEGIA: Ejecutar directamente fnModificarStock() en lugar de hacer click
    # Primero, intentar obtener el onclick attribute
    onclick_value = btn_existencias.get_attribute("onclick")
    
    # Extraer el ID del producto del onclick (ej: fnModificarStock(57072189))
    id_producto = None
    if onclick_value:
        import re
        match = re.search(r'fnModificarStock\((\d+)\)', onclick_value)
        if match:
            id_producto = match.group(1)
            log.info(f"  ✓ ID del producto extraído: {id_producto}")
        else:
            log.warning(f"  ✗ No se pudo extraer ID desde onclick: {onclick_value}")
    else:
        log.warning(f"  ✗ El atributo onclick es None o no existe")
    
    # Si se extrajo el ID, ejecutar directamente fnModificarStock
    if id_producto:
        log.info(f"  → Ejecutando fnModificarStock({id_producto}) directamente...")
        try:
            # Intentar ejecutar la función
            resultado = driver.execute_script(f"return fnModificarStock({id_producto});")
            log.info(f"  ✓ Función fnModificarStock({id_producto}) ejecutada correctamente")
        except Exception as e:
            log.error(f"  ✗ Error ejecutando fnModificarStock: {e}")
    else:
        # Si no se extrajo el ID, hacer click normal
        log.info("  → No se pudo extraer ID, haciendo click normal...")
        try:
            btn_existencias.click()
            log.info("  ✓ Click realizado")
        except Exception as e:
            log.error(f"  ✗ Error en click: {e}")
    
    log.info("  → Esperando que aparezca el modal...")
    time.sleep(3 + extra_espera)

    # 5. Esperar a que el formulario del modal aparezca de múltiples formas
    form_encontrado = False
    
    # Intentar 1: Buscar directamente por ID formStock
    try:
        WebDriverWait(driver, WAIT_CORTO).until(
            EC.presence_of_element_located((By.ID, "formStock"))
        )
        log.info("  Formulario formStock encontrado en DOM")
        form_encontrado = True
    except TimeoutException:
        log.warning("  Primer intento: formStock no encontrado")
    
    # Intentar 2: Buscar usando presencia de atributo action="/MejoraBasica/ModificarStock"
    if not form_encontrado:
        try:
            formulario = WebDriverWait(driver, WAIT_CORTO).until(
                EC.presence_of_element_located((
                    By.XPATH, 
                    "//form[contains(@action,'ModificarStock')]"
                ))
            )
            log.info("  Formulario encontrado por action=/MejoraBasica/ModificarStock")
            form_encontrado = True
        except TimeoutException:
            log.warning("  Segundo intento: formulario con action no encontrado")
    
    # Intentar 3: Buscar cualquier form que contenga un campo N_Stock
    if not form_encontrado:
        try:
            formulario = WebDriverWait(driver, WAIT_CORTO).until(
                EC.presence_of_element_located((
                    By.XPATH, 
                    "//form[.//input[@name='N_Stock']]"
                ))
            )
            log.info("  Formulario encontrado por input N_Stock")
            form_encontrado = True
        except TimeoutException:
            log.warning("  Tercer intento: formulario con input N_Stock no encontrado")
    
    # Capturar el HTML actual para debugging
    html_actual = driver.find_element(By.TAG_NAME, "body").get_attribute("innerHTML")
    if "formStock" in html_actual:
        log.info("  INFO: 'formStock' EXISTE en el HTML actual")
    else:
        log.warning("  INFO: 'formStock' NO existe en el HTML actual")
    
    if "N_Stock" in html_actual:
        log.info("  INFO: 'N_Stock' EXISTE en el HTML actual")
        # Mostrar contexto del N_Stock
        import re
        match = re.search(r'.{0,150}N_Stock.{0,150}', html_actual)
        if match:
            log.info(f"  Contexto N_Stock: ...{match.group(0)}...")
    else:
        log.warning("  INFO: 'N_Stock' NO existe en el HTML actual")
    
    time.sleep(1)

    # 6. Buscar el campo N_Stock 
    campo_stock = None
    try:
        campo_stock = driver.find_element(By.ID, "N_Stock")
        log.info("  Campo N_Stock encontrado por ID")
    except NoSuchElementException:
        log.warning("  Campo N_Stock no encontrado por ID, buscando en el formulario...")
        try:
            campo_stock = driver.find_element(By.CSS_SELECTOR, "#formStock #N_Stock")
            log.info("  Campo N_Stock encontrado en formulario")
        except NoSuchElementException:
            pass
    
    if campo_stock is None:
        # Último intento: buscar visible en el modal
        try:
            modal = driver.find_element(By.CSS_SELECTOR, ".modal.show, .modal.in, [style*='display: block']")
            inputs = modal.find_elements(By.CSS_SELECTOR, "input[id='N_Stock']")
            if inputs:
                campo_stock = inputs[0]
                log.info("  Campo N_Stock encontrado en modal visible")
        except:
            pass
    
    if campo_stock is None:
        # Mostrar el HTML del body para debugging
        try:
            body_html = driver.find_element(By.TAG_NAME, "body").get_attribute("innerHTML")
            if "N_Stock" in body_html:
                log.warning("  El campo N_Stock EXISTE en el HTML pero no se puede acceder")
            else:
                log.warning("  El campo N_Stock NO existe en el HTML actual")
        except:
            pass
        raise TimeoutException("No se encontro el campo de stock (N_Stock)")

    # 7. Limpiar el campo y escribir el nuevo stock
    log.info(f"  Limpiando campo y escribiendo stock: {stock}")
    campo_stock.clear()
    time.sleep(0.5)
    campo_stock.send_keys(stock)
    log.info(f"  Stock ingresado: {stock}")

    time.sleep(1)

    # 8. Buscar y hacer click en el boton de guardar
    btn_guardar = None
    selectores_guardar = [
        (By.ID, "btn_guardar"),
        (By.XPATH, "//button[@type='submit']"),
        (By.XPATH, "//button[contains(text(),'Guardar')]"),
    ]
    
    for selector_type, selector in selectores_guardar:
        try:
            btn_guardar = WebDriverWait(driver, WAIT_CORTO).until(
                EC.element_to_be_clickable((selector_type, selector))
            )
            log.info(f"  Boton guardar encontrado, haciendo click...")
            btn_guardar.click()
            log.info(f"  Click en guardar realizado")
            break
        except TimeoutException:
            continue

    if btn_guardar is None:
        log.warning("  No se encontro boton guardar. Intentando presionar Enter...")
        campo_stock.send_keys("\n")

    time.sleep(2)

    # 9. Manejar confirmacion "¿Esta seguro de registrar los cambios realizados?"
    confirmado = False

    # Intentar alert nativo primero
    try:
        alerta = WebDriverWait(driver, WAIT_CORTO).until(EC.alert_is_present())
        log.info(f"  Alerta detectada, aceptando...")
        alerta.accept()
        confirmado = True
    except TimeoutException:
        pass

    # Si no fue alert nativo, buscar el boton "Sí" del modal personalizado
    if not confirmado:
        log.info("  Buscando boton de confirmacion en modal...")
        selectores_si = [
            (By.XPATH, "//div[contains(@class,'_wModal_btn_ok')]"),
            (By.XPATH, "//div[contains(@class,'_wModal_btn') and contains(text(),'Sí')]"),
            (By.XPATH, "//button[contains(@class,'_wModal_btn_ok')]"),
            (By.XPATH, "//button[contains(text(),'Sí')]"),
        ]
        
        for selector_type, selector in selectores_si:
            try:
                btn_si = WebDriverWait(driver, WAIT_CORTO).until(
                    EC.element_to_be_clickable((selector_type, selector))
                )
                log.info(f"  Boton Sí encontrado, haciendo click...")
                btn_si.click()
                log.info(f"  Confirmacion aceptada")
                confirmado = True
                break
            except TimeoutException:
                continue

    if not confirmado:
        log.warning("  No se detecto dialogo de confirmacion")

    time.sleep(2)

    # 10. Cerrar cualquier modal abierto
    try:
        btn_cerrar = WebDriverWait(driver, WAIT_CORTO).until(
            EC.element_to_be_clickable((
                By.XPATH,
                "//button[contains(@class,'close')]"
                "|//button[@data-dismiss='modal']"
                "|//div[contains(@class,'_wModal_close')]"
            ))
        )
        btn_cerrar.click()
        log.info("  Modal cerrado")
    except TimeoutException:
        log.info("  No hay modal para cerrar")

    time.sleep(2)
    log.info(f"  Producto {parte} actualizado exitosamente")


def recuperar_estado(driver):
    """Intenta cerrar modales abiertos y limpiar la busqueda."""
    # Cerrar cualquier modal abierto
    try:
        modales_cerrar = driver.find_elements(
            By.XPATH,
            "//div[contains(@class,'modal')]//button[contains(@class,'close')]"
            "|//button[@data-dismiss='modal']",
        )
        for btn in modales_cerrar:
            try:
                btn.click()
                time.sleep(0.5)
            except Exception:
                pass
    except Exception:
        pass

    # Cerrar alertas pendientes
    try:
        driver.switch_to.alert.accept()
    except NoAlertPresentException:
        pass

    # Limpiar campo de busqueda
    try:
        campo = driver.find_element(By.ID, "C_Descripcion")
        campo.clear()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# APRENDIZAJE ADAPTATIVO
# ---------------------------------------------------------------------------
class AnalizadorFallos:
    """
    Registra patrones de error durante la ejección y activa ajustes automáticos
    para reducir la probabilidad de que el mismo fallo se repita.
    El historial acumulado se guarda en aprendizaje.json para persistir entre sesiones.
    """

    ARCHIVO = Path(__file__).parent / "aprendizaje.json"
    UMBRAL = 3  # fallos del mismo tipo necesarios para activar un ajuste

    def __init__(self):
        self.historial = {}       # tipo_fallo -> count (sesión actual)
        self.acumulado = {}       # tipo_fallo -> count (histórico)
        self.ajustes_activos = set()
        self._cargar()
        # Activar ajustes que ya superaron el umbral en sesiones anteriores
        for tipo, cnt in self.acumulado.items():
            if cnt >= self.UMBRAL:
                self.ajustes_activos.add(tipo)
        if self.ajustes_activos:
            log.info(
                f"[APRENDIZAJE] Ajustes cargados de sesión anterior: "
                f"{', '.join(self.ajustes_activos)}"
            )

    def _cargar(self):
        try:
            if self.ARCHIVO.exists():
                data = json.loads(self.ARCHIVO.read_text(encoding="utf-8"))
                self.acumulado = data.get("acumulado", {})
        except Exception:
            self.acumulado = {}

    def registrar(self, tipo_fallo: str):
        """Registra un fallo y activa ajustes si se supera el umbral."""
        if not tipo_fallo:
            return
        self.historial[tipo_fallo] = self.historial.get(tipo_fallo, 0) + 1
        self.acumulado[tipo_fallo] = self.acumulado.get(tipo_fallo, 0) + 1
        count = self.historial[tipo_fallo]
        if count >= self.UMBRAL and tipo_fallo not in self.ajustes_activos:
            self.ajustes_activos.add(tipo_fallo)
            log.info(
                f"[APRENDIZAJE] Patrón detectado: '{tipo_fallo}' ocurrió {count} veces "
                f"→ ajuste activado para el resto de la ejecución"
            )

    def wait_extra(self) -> float:
        """Segundos de espera adicionales si hay timeouts recurrentes."""
        return 5.0 if "Tiempo de espera agotado" in self.ajustes_activos else 0.0

    def forzar_recarga(self) -> bool:
        """True si debe recargar filtros antes de cada búsqueda."""
        return "Producto no encontrado en la tabla" in self.ajustes_activos

    def preferir_js(self) -> bool:
        """True si debe preferir ejecución JS en vez de clic para abrir el modal."""
        return "Modal de stock no se abrió" in self.ajustes_activos

    def resumen(self) -> str:
        if not self.ajustes_activos:
            return "Sin ajustes activos"
        desc = {
            "Tiempo de espera agotado": "+5s de espera extra por producto",
            "Producto no encontrado en la tabla": "recarga de filtros antes de buscar",
            "Modal de stock no se abrió": "ejecución JS directa para modal",
        }
        return " | ".join(desc.get(a, a) for a in self.ajustes_activos)

    def guardar(self):
        try:
            data = {
                "acumulado": self.acumulado,
                "ultima_sesion": str(datetime.now()),
                "nota": "Generado automáticamente. Borra este archivo para resetear el aprendizaje.",
            }
            self.ARCHIVO.write_text(
                json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            log.info(f"[APRENDIZAJE] Conocimiento guardado en: {self.ARCHIVO.name}")
        except Exception as e:
            log.warning(f"[APRENDIZAJE] No se pudo guardar: {e}")


# ---------------------------------------------------------------------------
# EJECUCION REUTILIZABLE (CLI + GUI)
# ---------------------------------------------------------------------------
def ejecutar_bot(
    excel_path: Path,
    acuerdo_texto: str,
    catalogo_texto: str,
    categoria_texto: str,
    pausa_entre_productos: int = 2,
):
    """Ejecuta el flujo completo del bot con configuración dinámica."""
    global EXCEL_PATH, ACUERDO_TEXTO, CATALOGO_TEXTO, CATEGORIA_TEXTO
    global PAUSA_ENTRE_PRODUCTOS, REPORTE_PATH, RESULTADOS
    global PAUSA_EVENTO, DETENER_EVENTO, ANALIZADOR
    RESULTADOS = []
    # Inicializar eventos para modo CLI (la GUI los crea en _worker_run antes de llamar acá)
    if PAUSA_EVENTO is None:
        PAUSA_EVENTO = threading.Event()
        PAUSA_EVENTO.set()
    if DETENER_EVENTO is None:
        DETENER_EVENTO = threading.Event()
    ANALIZADOR = AnalizadorFallos()
    log.info(f"[APRENDIZAJE] Ajustes al inicio: {ANALIZADOR.resumen()}") 

    EXCEL_PATH = Path(excel_path)
    ACUERDO_TEXTO = acuerdo_texto
    CATALOGO_TEXTO = catalogo_texto
    CATEGORIA_TEXTO = categoria_texto
    PAUSA_ENTRE_PRODUCTOS = pausa_entre_productos
    REPORTE_PATH = nueva_ruta_reporte()

    log.info("Iniciando Peru Compras Bot...")
    log.info(f"Archivo Excel seleccionado: {EXCEL_PATH}")
    log.info(f"Filtro acuerdo: {ACUERDO_TEXTO}")
    log.info(f"Filtro catálogo: {CATALOGO_TEXTO}")
    log.info(f"Filtro categoría: {CATEGORIA_TEXTO}")

    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"No se encontro el archivo Excel: {EXCEL_PATH}")

    df = pd.read_excel(EXCEL_PATH)
    if "Parte" not in df.columns or "Stock" not in df.columns:
        raise ValueError("El Excel debe tener columnas 'Parte' y 'Stock'.")

    df = df.dropna(subset=["Parte", "Stock"]).reset_index(drop=True)
    log.info(f"Productos cargados: {len(df)}")
    if len(df) == 0:
        raise ValueError("No hay productos para procesar en el Excel.")

    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)

    try:
        paso1_login(driver)
        paso2_navegacion(driver)
        paso3_filtros(driver)
        paso4_actualizar_stock(driver, df)
        generar_reporte_excel(acuerdo_texto, catalogo_texto, categoria_texto)
        if ANALIZADOR:
            ANALIZADOR.guardar()
        return REPORTE_PATH
    finally:
        driver.quit()
        log.info("Navegador cerrado. Fin del script.")


def main_cli():
    """Modo clásico por consola."""
    global MODO_GUI, EVENTO_LOGIN, GUI_NOTIFICAR_LOGIN
    MODO_GUI = False
    EVENTO_LOGIN = None
    GUI_NOTIFICAR_LOGIN = None

    try:
        reporte = ejecutar_bot(
            excel_path=EXCEL_PATH,
            acuerdo_texto=ACUERDO_TEXTO,
            catalogo_texto=CATALOGO_TEXTO,
            categoria_texto=CATEGORIA_TEXTO,
            pausa_entre_productos=PAUSA_ENTRE_PRODUCTOS,
        )
        print("\n" + "=" * 60)
        print(f"Reporte guardado en: {reporte}")
        input("Presiona ENTER para cerrar...")
    except KeyboardInterrupt:
        log.info("\nEjecucion interrumpida por el usuario.")
    except Exception as e:
        log.error(f"Error fatal: {e}", exc_info=True)


class TextQueueLogHandler(logging.Handler):
    def __init__(self, cola: Queue):
        super().__init__()
        self.cola = cola

    def emit(self, record):
        try:
            msg = self.format(record)
            self.cola.put(msg)
        except Exception:
            pass


class _Tooltip:
    """Tooltip sencillo que aparece al pasar el mouse sobre un widget."""
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self._tip = None
        widget.bind("<Enter>", self._show)
        widget.bind("<Leave>", self._hide)
        widget.bind("<ButtonPress>", self._hide)

    def _show(self, _event=None):
        x, y, _, cy = self.widget.bbox("insert") if hasattr(self.widget, "bbox") else (0, 0, 0, 0)
        x += self.widget.winfo_rootx() + 24
        y += self.widget.winfo_rooty() + cy + 20
        self._tip = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        lbl = tk.Label(
            tw, text=self.text, justify="left",
            background="#FFFBE6", foreground="#333333",
            relief="flat", borderwidth=1,
            font=("Segoe UI", 8),
            wraplength=320, padx=6, pady=4,
        )
        lbl.pack()

    def _hide(self, _event=None):
        if self._tip:
            self._tip.destroy()
            self._tip = None


class PeruComprasGUI:
    # ── Paleta de colores ──────────────────────────────────────────────
    C_FONDO       = "#F4F6FB"
    C_HEADER      = "#00205B"
    C_STEP_ACTIVE = "#1A73E8"
    C_STEP_DONE   = "#34A853"
    C_STEP_IDLE   = "#9AA0A6"
    C_ACCION      = "#1A73E8"
    C_PELIGRO     = "#D93025"
    C_ADVERTENCIA = "#F29900"
    C_TEXTO       = "#202124"
    C_TEXTO_SUAVE = "#5F6368"
    C_BORDE       = "#DADCE0"
    C_LOGIN_BG    = "#FFF8E1"
    C_LOGIN_BORDE = "#F29900"

    def __init__(self, root):
        self.root = root
        self.root.configure(bg=self.C_FONDO)

        self.log_queue = Queue()
        self.login_event = threading.Event()
        self.worker = None
        self.reporte_generado = None
        self._pausado = False
        self._total_productos = 0
        self._procesados = 0

        self.excel_var    = tk.StringVar(value=str(Path(__file__).parent / "productos.xlsx"))
        self.acuerdo_var  = tk.StringVar(value=ACUERDO_TEXTO)
        self.catalogo_var = tk.StringVar(value=CATALOGO_TEXTO)
        self.categoria_var = tk.StringVar(value=CATEGORIA_TEXTO)
        self.pausa_var    = tk.StringVar(value=str(PAUSA_ENTRE_PRODUCTOS))
        self.estado_var   = tk.StringVar(value="")

        self._apply_theme()
        self._build_ui()
        self._configurar_logging_gui()
        self._tick_logs()

    # ------------------------------------------------------------------
    # Tema ttk personalizado
    # ------------------------------------------------------------------
    def _apply_theme(self):
        style = ttk.Style(self.root)
        # Usar 'clam' como base (compatible con Windows/Linux)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure(".", font=("Segoe UI", 9), background=self.C_FONDO, foreground=self.C_TEXTO)
        style.configure("TFrame",      background=self.C_FONDO)
        style.configure("TLabel",      background=self.C_FONDO, foreground=self.C_TEXTO)
        style.configure("TLabelframe", background=self.C_FONDO)
        style.configure("TLabelframe.Label", background=self.C_FONDO, foreground=self.C_HEADER,
                        font=("Segoe UI", 9, "bold"))
        style.configure("TEntry",      fieldbackground="#FFFFFF", bordercolor=self.C_BORDE)
        style.configure("TCombobox",   fieldbackground="#FFFFFF", bordercolor=self.C_BORDE)

        # Botón principal (azul)
        style.configure("Accion.TButton",
                        background=self.C_ACCION, foreground="#FFFFFF",
                        font=("Segoe UI", 10, "bold"),
                        padding=(14, 8), relief="flat", borderwidth=0)
        style.map("Accion.TButton",
                  background=[("active", "#1557B0"), ("disabled", "#BDC1C6")],
                  foreground=[("disabled", "#FFFFFF")])

        # Botón secundario (gris)
        style.configure("Secundario.TButton",
                        background="#FFFFFF", foreground=self.C_TEXTO,
                        font=("Segoe UI", 9),
                        padding=(10, 6), relief="flat", borderwidth=1)
        style.map("Secundario.TButton",
                  background=[("active", "#E8EAED")])

        # Botón peligro (rojo, detener)
        style.configure("Peligro.TButton",
                        background="#FDECEA", foreground=self.C_PELIGRO,
                        font=("Segoe UI", 9, "bold"),
                        padding=(10, 6), relief="flat", borderwidth=0)
        style.map("Peligro.TButton",
                  background=[("active", "#FAD2CF"), ("disabled", "#F5F5F5")],
                  foreground=[("disabled", "#9AA0A6")])

        # Botón advertencia (amarillo, pausa)
        style.configure("Pausa.TButton",
                        background="#FFF8E1", foreground=self.C_ADVERTENCIA,
                        font=("Segoe UI", 9, "bold"),
                        padding=(10, 6), relief="flat", borderwidth=0)
        style.map("Pausa.TButton",
                  background=[("active", "#FFF3CD"), ("disabled", "#F5F5F5")],
                  foreground=[("disabled", "#9AA0A6")])

        # Botón login (llamativo)
        style.configure("Login.TButton",
                        background=self.C_LOGIN_BG, foreground="#33691E",
                        font=("Segoe UI", 10, "bold"),
                        padding=(14, 10), relief="flat", borderwidth=2)
        style.map("Login.TButton",
                  background=[("active", "#F9FBE7"), ("disabled", "#F5F5F5")],
                  foreground=[("disabled", "#9AA0A6")])

        # Barra de progreso
        style.configure("Verde.Horizontal.TProgressbar",
                        troughcolor=self.C_BORDE,
                        background=self.C_STEP_DONE,
                        thickness=14)

    # ------------------------------------------------------------------
    # Construcción de la UI
    # ------------------------------------------------------------------
    def _build_ui(self):
        # ── Cabecera ──────────────────────────────────────────────────
        header = tk.Frame(self.root, bg=self.C_HEADER, height=56)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Label(
            header,
            text="  Peru Compras Bot",
            bg=self.C_HEADER, fg="#FFFFFF",
            font=("Segoe UI", 15, "bold"),
            anchor="w",
        ).pack(side="left", padx=16, pady=10)

        tk.Label(
            header,
            text="Automatización de actualización de stock",
            bg=self.C_HEADER, fg="#A8C7FA",
            font=("Segoe UI", 9),
        ).pack(side="left", padx=(0, 16), pady=18)

        # ── Banda de estado (banner) ───────────────────────────────────
        self._banner_frame = tk.Frame(self.root, bg="#E8F0FE", height=32)
        self._banner_frame.pack(fill="x")
        self._banner_frame.pack_propagate(False)
        self._banner_lbl = tk.Label(
            self._banner_frame, textvariable=self.estado_var,
            bg="#E8F0FE", fg=self.C_STEP_ACTIVE,
            font=("Segoe UI", 9, "italic"), anchor="w",
        )
        self._banner_lbl.pack(side="left", padx=16, pady=6)

        # ── Contenedor principal con scroll ───────────────────────────
        scroll_outer = tk.Frame(self.root, bg=self.C_FONDO)
        scroll_outer.pack(fill="both", expand=True)

        canvas = tk.Canvas(scroll_outer, bg=self.C_FONDO, highlightthickness=0)
        scrollbar = ttk.Scrollbar(scroll_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self._scroll_frame = tk.Frame(canvas, bg=self.C_FONDO)
        self._scroll_win = canvas.create_window((0, 0), window=self._scroll_frame, anchor="nw")

        def _on_resize(e):
            canvas.itemconfig(self._scroll_win, width=e.width)
        canvas.bind("<Configure>", _on_resize)

        def _on_frame_configure(_e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        self._scroll_frame.bind("<Configure>", _on_frame_configure)

        def _mousewheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _mousewheel)

        contenedor = self._scroll_frame

        # ──────────────────────────────────────────────────────────────
        # Helper local: crea un card numerado y devuelve el frame de contenido
        #──────────────────────────────────────────────────────────────
        def _card(num: str, title: str) -> tk.Frame:
            outer = tk.Frame(contenedor, bg=self.C_FONDO)
            outer.pack(fill="x", padx=20, pady=(0, 12))
            tk.Label(
                outer, text=num,
                bg=self.C_STEP_ACTIVE, fg="#FFFFFF",
                font=("Segoe UI", 10, "bold"), width=2, relief="flat",
            ).pack(side="left", anchor="n", padx=(0, 10), pady=4)
            card = tk.Frame(
                outer, bg="#FFFFFF",
                highlightbackground=self.C_BORDE, highlightthickness=1,
            )
            card.pack(side="left", fill="both", expand=True)
            tk.Label(
                card,
                text=f"  Paso {num}  —  {title}",
                bg=self.C_HEADER, fg="#FFFFFF",
                font=("Segoe UI", 9, "bold"),
                anchor="w", padx=10, pady=7,
            ).pack(fill="x")
            content = tk.Frame(card, bg="#FFFFFF")
            content.pack(fill="x", padx=14, pady=10)
            return content

        # ════════════════════════════════════════════════════════════════
        # PASO 1 — Archivo Excel
        # ════════════════════════════════════════════════════════════════
        f1 = _card("1", "Selecciona tu archivo de productos")

        # Sub-panel: ¿Tienes plantilla?
        aviso_plt = tk.Frame(f1, bg="#E8F5E9", bd=0, highlightbackground="#A5D6A7", highlightthickness=1)
        aviso_plt.pack(fill="x", pady=(0, 10))
        tk.Label(
            aviso_plt,
            text="  ¿Primera vez? Descarga la plantilla de ejemplo, rellénala y luego selecciónala aquí.",
            bg="#E8F5E9", fg="#2E7D32",
            font=("Segoe UI", 8),
            anchor="w",
        ).pack(side="left", padx=6, pady=6)
        btn_plt = ttk.Button(aviso_plt, text="⬇ Descargar plantilla",
                             command=self._descargar_plantilla, style="Secundario.TButton")
        btn_plt.pack(side="right", padx=8, pady=4)
        _Tooltip(btn_plt, "Descarga un Excel de ejemplo con las columnas correctas\ny llénalo con tus productos.")

        # Fila de selección
        fila_excel = ttk.Frame(f1)
        fila_excel.pack(fill="x")
        fila_excel.columnconfigure(0, weight=1)

        self.entry_excel = ttk.Entry(fila_excel, textvariable=self.excel_var, font=("Segoe UI", 9))
        self.entry_excel.grid(row=0, column=0, sticky="we", padx=(0, 8))
        btn_sel = ttk.Button(fila_excel, text="📂 Seleccionar archivo...",
                             command=self._seleccionar_excel, style="Secundario.TButton")
        btn_sel.grid(row=0, column=1)
        _Tooltip(btn_sel, "Busca y selecciona tu archivo Excel (.xlsx)\ncon las columnas 'Parte' y 'Stock'.")

        ttk.Label(
            f1,
            text="El archivo debe tener dos columnas: 'Parte' (código del producto) y 'Stock' (cantidad).",
            foreground=self.C_TEXTO_SUAVE, font=("Segoe UI", 8),
        ).pack(anchor="w", pady=(6, 0))

        # ════════════════════════════════════════════════════════════════
        # PASO 2 — Filtros del portal
        # ════════════════════════════════════════════════════════════════
        f2 = _card("2", "Configura los filtros del portal")

        # Aviso "cargar desde el portal"
        aviso_f = tk.Frame(f2, bg="#E3F2FD", bd=0, highlightbackground="#90CAF9", highlightthickness=1)
        aviso_f.pack(fill="x", pady=(0, 10))
        tk.Label(
            aviso_f,
            text="  Si los desplegables están vacíos, usa el botón para traer las opciones directamente del portal.",
            bg="#E3F2FD", fg="#1565C0",
            font=("Segoe UI", 8), anchor="w",
        ).pack(side="left", padx=6, pady=6)
        self.btn_cargar_opts = ttk.Button(
            aviso_f, text="🔄 Importar opciones del portal",
            command=self._cargar_opciones, style="Secundario.TButton",
        )
        self.btn_cargar_opts.pack(side="right", padx=8, pady=4)
        _Tooltip(
            self.btn_cargar_opts,
            "Abre Chrome, te pedirá que inicies sesión y luego\n"
            "cargará automáticamente los Acuerdos, Catálogos\n"
            "y Categorías disponibles en el portal.",
        )

        grid_f = tk.Frame(f2, bg="#FFFFFF")
        grid_f.pack(fill="x")
        grid_f.columnconfigure(1, weight=1)

        labels_filtros = ["Acuerdo Marco:", "Catálogo:", "Categoría:"]
        tips_filtros = [
            "Selecciona el Acuerdo Marco al que pertenecen tus productos.\nEjemplo: EXT-CE-2022-5 COMPUTADORAS DE ESCRITORIO",
            "Selecciona el Catálogo Electrónico correspondiente.\nEjemplo: COMPUTADORAS DE ESCRITORIO",
            "Selecciona la Categoría específica dentro del catálogo.\nEjemplo: MONITOR",
        ]
        self.combo_acuerdo  = self._make_combo_row(grid_f, 0, labels_filtros[0], self.acuerdo_var,  tips_filtros[0])
        self.combo_catalogo = self._make_combo_row(grid_f, 1, labels_filtros[1], self.catalogo_var, tips_filtros[1])
        self.combo_categoria = self._make_combo_row(grid_f, 2, labels_filtros[2], self.categoria_var, tips_filtros[2])
        self.combo_acuerdo.bind("<<ComboboxSelected>>",  self._on_acuerdo_changed)
        self.combo_catalogo.bind("<<ComboboxSelected>>", self._on_catalogo_changed)

        # Configuración avanzada (colapsable)
        self._avanzado_visible = tk.BooleanVar(value=False)
        btn_avanzado = ttk.Button(
            f2, text="⚙ Configuración avanzada ▸",
            command=self._toggle_avanzado, style="Secundario.TButton",
        )
        btn_avanzado.pack(anchor="w", pady=(8, 0))
        self._btn_avanzado = btn_avanzado

        self._frame_avanzado = tk.Frame(f2, bg="#FFFFFF")
        # (No se hace pack aquí; se muestra solo si el usuario lo abre)
        av_lbl = ttk.Label(self._frame_avanzado, text="Pausa entre productos (segundos):",
                           foreground=self.C_TEXTO_SUAVE, font=("Segoe UI", 8))
        av_lbl.pack(side="left", pady=(4, 0))
        av_entry = ttk.Entry(self._frame_avanzado, textvariable=self.pausa_var, width=6)
        av_entry.pack(side="left", padx=8, pady=(4, 0))
        _Tooltip(av_entry,
                 "Tiempo de espera entre cada producto (en segundos).\n"
                 "Valor por defecto: 2. Auméntalo si el portal responde lento.")

        # ════════════════════════════════════════════════════════════════
        # PASO 3 — Ejecutar
        # ════════════════════════════════════════════════════════════════
        f3 = _card("3", "Inicia la automatización")

        # Botón principal de inicio
        self.btn_iniciar = ttk.Button(
            f3, text="▶  Iniciar actualización de stock",
            command=self._iniciar, style="Accion.TButton",
        )
        self.btn_iniciar.pack(fill="x", ipady=4, pady=(0, 10))
        _Tooltip(self.btn_iniciar,
                 "Abre Chrome, te pedirá que inicies sesión\n"
                 "y luego actualizará el stock de todos los productos de tu Excel.")

        # Panel de login (visible solo cuando Chrome espera login)
        self._panel_login = tk.Frame(
            f3, bg=self.C_LOGIN_BG, bd=0,
            highlightbackground=self.C_LOGIN_BORDE, highlightthickness=2,
        )
        # No se empaca aún – aparece solo cuando se necesita
        tk.Label(
            self._panel_login,
            text="⏳  El bot está esperando que inicies sesión en Chrome",
            bg=self.C_LOGIN_BG, fg="#7B5800",
            font=("Segoe UI", 10, "bold"), anchor="w",
        ).pack(anchor="w", padx=14, pady=(10, 4))
        tk.Label(
            self._panel_login,
            text="1. Ve a la ventana de Chrome que se abrió.\n"
                 "2. Escribe tu RUC, usuario y contraseña.\n"
                 "3. Haz clic en 'Ingresar'.\n"
                 "4. Regresa aquí y haz clic en el botón verde de abajo.",
            bg=self.C_LOGIN_BG, fg="#5D4037",
            font=("Segoe UI", 9), anchor="w", justify="left",
        ).pack(anchor="w", padx=14, pady=(0, 8))
        self.btn_login = ttk.Button(
            self._panel_login,
            text="✅  Ya inicié sesión — continuar",
            command=self._continuar_login,
            style="Login.TButton",
        )
        self.btn_login.pack(fill="x", padx=14, pady=(0, 12), ipady=2)

        # Panel de controles de ejecución (visible durante proceso)
        self._panel_ctrl = tk.Frame(f3, bg="#FFFFFF")
        # No se empaca aún

        # Barra de progreso
        prog_frame = tk.Frame(self._panel_ctrl, bg="#FFFFFF")
        prog_frame.pack(fill="x", pady=(0, 6))
        self._lbl_progreso = tk.Label(
            prog_frame, text="Preparando...",
            bg="#FFFFFF", fg=self.C_TEXTO_SUAVE, font=("Segoe UI", 8),
        )
        self._lbl_progreso.pack(anchor="w")
        self.progress = ttk.Progressbar(
            prog_frame, orient="horizontal", mode="determinate",
            style="Verde.Horizontal.TProgressbar",
        )
        self.progress.pack(fill="x", pady=(2, 0))

        # Botones Pausar / Detener
        ctrl_btns = tk.Frame(self._panel_ctrl, bg="#FFFFFF")
        ctrl_btns.pack(fill="x", pady=(6, 0))
        self.btn_pausar = ttk.Button(
            ctrl_btns, text="⏸  Pausar",
            command=self._pausar_reanudar, style="Pausa.TButton",
        )
        self.btn_pausar.pack(side="left", padx=(0, 8))
        _Tooltip(self.btn_pausar, "Pausa el proceso después de que termine el producto actual.\nPuedes reanudarlo cuando quieras.")

        self.btn_detener = ttk.Button(
            ctrl_btns, text="⏹  Detener y generar reporte",
            command=self._detener, style="Peligro.TButton",
        )
        self.btn_detener.pack(side="left")
        _Tooltip(self.btn_detener, "Detiene el proceso y genera el reporte Excel\ncon los resultados hasta el momento.")

        # ════════════════════════════════════════════════════════════════
        # PANEL DE RESULTADO (visible cuando termina)
        # ════════════════════════════════════════════════════════════════
        self._panel_resultado = tk.Frame(
            contenedor, bg="#E8F5E9", bd=0,
            highlightbackground="#A5D6A7", highlightthickness=2,
        )
        # No se empaca aún

        tk.Label(
            self._panel_resultado,
            text="✅  ¡Proceso completado exitosamente!",
            bg="#E8F5E9", fg="#1B5E20",
            font=("Segoe UI", 11, "bold"),
        ).pack(anchor="w", padx=14, pady=(12, 4))

        self._lbl_resultado_info = tk.Label(
            self._panel_resultado,
            text="", bg="#E8F5E9", fg="#2E7D32",
            font=("Segoe UI", 9), anchor="w", justify="left",
        )
        self._lbl_resultado_info.pack(anchor="w", padx=14, pady=(0, 6))

        res_btns = tk.Frame(self._panel_resultado, bg="#E8F5E9")
        res_btns.pack(anchor="w", padx=14, pady=(0, 12))
        self.btn_abrir_reporte = ttk.Button(
            res_btns, text="📊  Abrir reporte Excel",
            command=self._abrir_reporte, style="Accion.TButton",
        )
        self.btn_abrir_reporte.pack(side="left", padx=(0, 10))
        ttk.Button(
            res_btns, text="📁  Abrir carpeta",
            command=self._abrir_carpeta, style="Secundario.TButton",
        ).pack(side="left")

        # ════════════════════════════════════════════════════════════════
        # LOG — actividad en tiempo real (colapsable)
        # ════════════════════════════════════════════════════════════════
        log_header = tk.Frame(contenedor, bg=self.C_FONDO)
        log_header.pack(fill="x", padx=20, pady=(0, 4))

        self._log_visible = tk.BooleanVar(value=False)
        self._btn_toggle_log = ttk.Button(
            log_header, text="📋  Ver actividad detallada ▸",
            command=self._toggle_log, style="Secundario.TButton",
        )
        self._btn_toggle_log.pack(side="left")
        _Tooltip(self._btn_toggle_log,
                 "Muestra/oculta el registro técnico en tiempo real.\n"
                 "Útil para ver qué está haciendo el bot en cada momento.")

        ttk.Button(
            log_header, text="📈  Estadísticas de errores",
            command=self._ver_aprendizaje, style="Secundario.TButton",
        ).pack(side="left", padx=(10, 0))

        self._frame_log = ttk.LabelFrame(contenedor, text="Actividad en tiempo real", padding=8)
        # Oculto por defecto

        self.txt_log = scrolledtext.ScrolledText(
            self._frame_log, height=16, wrap="word", state="disabled",
            font=("Consolas", 8), background="#1E1E1E", foreground="#D4D4D4",
            insertbackground="#FFFFFF",
        )
        self.txt_log.pack(fill="both", expand=True)

        # Colorear líneas del log
        self.txt_log.tag_configure("error",   foreground="#F48771")
        self.txt_log.tag_configure("warning", foreground="#FFD700")
        self.txt_log.tag_configure("ok",      foreground="#89D185")
        self.txt_log.tag_configure("paso",    foreground="#9CDCFE", font=("Consolas", 8, "bold"))

        # ── Pie de página ──────────────────────────────────────────────
        tk.Frame(contenedor, bg=self.C_BORDE, height=1).pack(fill="x", padx=20, pady=(14, 0))
        tk.Label(
            contenedor,
            text="Peru Compras Bot  •  Uso exclusivo interno",
            bg=self.C_FONDO, fg=self.C_TEXTO_SUAVE,
            font=("Segoe UI", 7),
        ).pack(pady=(4, 12))

    # ------------------------------------------------------------------
    # Helpers de construcción
    # ------------------------------------------------------------------
    # (_make_step_frame removed — step cards are built inline in _build_ui via _card())

    def _make_combo_row(self, parent, row: int, label: str, variable: tk.StringVar, tip: str) -> ttk.Combobox:
        tk.Label(parent, text=label, bg="#FFFFFF",
                 font=("Segoe UI", 9), fg=self.C_TEXTO).grid(
            row=row, column=0, sticky="w", padx=(0, 10), pady=4,
        )
        combo = ttk.Combobox(parent, textvariable=variable, state="normal")
        combo.grid(row=row, column=1, sticky="we", pady=4)
        _Tooltip(combo, tip)
        return combo

    def _toggle_avanzado(self):
        if self._avanzado_visible.get():
            self._frame_avanzado.pack_forget()
            self._btn_avanzado.configure(text="⚙ Configuración avanzada ▸")
            self._avanzado_visible.set(False)
        else:
            self._frame_avanzado.pack(anchor="w", pady=(4, 0))
            self._btn_avanzado.configure(text="⚙ Configuración avanzada ▾")
            self._avanzado_visible.set(True)

    def _toggle_log(self):
        if self._log_visible.get():
            self._frame_log.pack_forget()
            self._btn_toggle_log.configure(text="📋  Ver actividad detallada ▸")
            self._log_visible.set(False)
        else:
            self._frame_log.pack(fill="both", expand=True, padx=20, pady=(0, 10))
            self._btn_toggle_log.configure(text="📋  Ocultar actividad detallada ▾")
            self._log_visible.set(True)

    # ------------------------------------------------------------------
    # Actualización de estado visual
    # ------------------------------------------------------------------
    def _set_banner(self, msg: str, color_bg: str = "#E8F0FE", color_fg: str = None):
        self._banner_frame.configure(bg=color_bg)
        self._banner_lbl.configure(bg=color_bg, fg=color_fg or self.C_STEP_ACTIVE)
        self.estado_var.set(f"  {msg}")

    def _mostrar_panel_login(self, mostrar: bool):
        if mostrar:
            self._panel_login.pack(fill="x", pady=(0, 8))
        else:
            self._panel_login.pack_forget()

    def _mostrar_panel_ctrl(self, mostrar: bool):
        if mostrar:
            self._panel_ctrl.pack(fill="x")
        else:
            self._panel_ctrl.pack_forget()

    def _mostrar_panel_resultado(self, mostrar: bool):
        if mostrar:
            self._panel_resultado.pack(fill="x", padx=20, pady=(0, 14))
        else:
            self._panel_resultado.pack_forget()

    def _actualizar_progreso(self, procesados: int, total: int, estado_txt: str = ""):
        if total > 0:
            self.progress["maximum"] = total
            self.progress["value"] = procesados
            pct = int(procesados / total * 100)
            self._lbl_progreso.configure(
                text=f"Producto {procesados} de {total}  ({pct}%)  {estado_txt}"
            )
        else:
            self.progress["value"] = 0
            self._lbl_progreso.configure(text=estado_txt or "Preparando...")

    # ------------------------------------------------------------------
    # Logging con color
    # ------------------------------------------------------------------
    def _configurar_logging_gui(self):
        self.gui_handler = TextQueueLogHandler(self.log_queue)
        self.gui_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        log.addHandler(self.gui_handler)

    def _tick_logs(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                self.txt_log.configure(state="normal")
                msg_lower = msg.lower()
                if "[error]" in msg_lower:
                    tag = "error"
                elif "[warning]" in msg_lower:
                    tag = "warning"
                elif "[ok]" in msg_lower or "exito" in msg_lower or "exitosamente" in msg_lower:
                    tag = "ok"
                elif "paso" in msg_lower or "=" * 10 in msg:
                    tag = "paso"
                else:
                    tag = ""
                self.txt_log.insert("end", msg + "\n", tag)
                self.txt_log.see("end")
                self.txt_log.configure(state="disabled")

                # Extraer progreso del mensaje de log
                # Formato esperado: "--- Producto X/Y: ..."
                import re as _re
                m = _re.search(r"Producto (\d+)/(\d+)", msg)
                if m:
                    proc, tot = int(m.group(1)), int(m.group(2))
                    self._procesados = proc
                    self._total_productos = tot
                    self.root.after(0, lambda p=proc, t=tot: self._actualizar_progreso(p, t))

                if "PASO" in msg:
                    txt = msg.split("] ", 1)[-1]
                    self._set_banner(txt)

                # Detectar login completado (para ocultar panel login)
                if "login confirmado" in msg_lower or "login completado" in msg_lower:
                    self.root.after(0, lambda: self._mostrar_panel_login(False))

        except Empty:
            pass
        self.root.after(150, self._tick_logs)

    def _seleccionar_excel(self):
        ruta = filedialog.askopenfilename(
            title="Selecciona el archivo Excel",
            filetypes=[("Excel", "*.xlsx *.xls")],
            initialdir=str(Path(__file__).parent),
        )
        if ruta:
            self.excel_var.set(ruta)
            self._set_banner(f"Archivo seleccionado: {Path(ruta).name}", "#E8F0FE")

    def _descargar_plantilla(self):
        destino = filedialog.asksaveasfilename(
            title="Guardar plantilla como…",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="plantilla_productos.xlsx",
            initialdir=str(Path(__file__).parent),
        )
        if not destino:
            return
        try:
            generar_plantilla_excel(Path(destino))
            if messagebox.askyesno(
                "Plantilla creada",
                f"Plantilla guardada en:\n{destino}\n\n"
                "¿Quieres abrir el archivo ahora para revisarlo?",
            ):
                os.startfile(destino)
            self.excel_var.set(destino)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear la plantilla:\n{e}")

    def _abrir_carpeta(self):
        os.startfile(str(Path(__file__).parent))

    def _abrir_reporte(self):
        if self.reporte_generado and Path(self.reporte_generado).exists():
            os.startfile(str(self.reporte_generado))
            return
        messagebox.showinfo("Reporte", "Aún no hay reporte generado en esta sesión.")

    def _continuar_login(self):
        self.login_event.set()
        self.root.after(0, lambda: self._mostrar_panel_login(False))
        self._set_banner("Login confirmado — el bot continúa...", "#E8F5E9", "#1B5E20")

    def _notificar_login_ui(self):
        def _update():
            self._mostrar_panel_login(True)
            self._set_banner(
                "Inicia sesión en Chrome y luego haz clic en el botón verde",
                self.C_LOGIN_BG, "#7B5800",
            )
        self.root.after(0, _update)

    # ------------------------------------------------------------------
    # Cascada de filtros
    # ------------------------------------------------------------------
    def _on_acuerdo_changed(self, event=None):
        self.combo_catalogo["values"] = []
        self.catalogo_var.set("")
        self.combo_categoria["values"] = []
        self.categoria_var.set("")
        self._set_banner("Acuerdo cambiado — usa '🔄 Importar opciones del portal' para actualizar los filtros")

    def _on_catalogo_changed(self, event=None):
        self.combo_categoria["values"] = []
        self.categoria_var.set("")
        self._set_banner("Catálogo cambiado — usa '🔄 Importar opciones del portal' para actualizar categorías")

    # ------------------------------------------------------------------
    # Cargar opciones desde el portal
    # ------------------------------------------------------------------
    def _cargar_opciones(self):
        if self.worker and self.worker.is_alive():
            messagebox.showwarning("Ocupado", "Espera a que termine el proceso en curso.")
            return
        self.login_event.clear()
        self.btn_cargar_opts.configure(state="disabled")
        self.btn_iniciar.configure(state="disabled")
        self._set_banner("Abriendo Chrome para conectarse al portal...", "#E3F2FD", "#1565C0")
        self.worker = threading.Thread(target=self._cargar_opciones_worker, daemon=True)
        self.worker.start()

    def _cargar_opciones_worker(self):
        global MODO_GUI, EVENTO_LOGIN, GUI_NOTIFICAR_LOGIN
        MODO_GUI = True
        EVENTO_LOGIN = self.login_event
        GUI_NOTIFICAR_LOGIN = self._notificar_login_ui
        driver = None
        try:
            chrome_opts = Options()
            chrome_opts.add_argument("--start-maximized")
            chrome_opts.add_argument("--disable-blink-features=AutomationControlled")
            chrome_opts.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_opts.add_experimental_option("useAutomationExtension", False)
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_opts)

            paso1_login(driver)
            paso2_navegacion(driver)

            # --- Acuerdo Marco ---
            acuerdo_opts = leer_opciones_select(driver, "ajaxAcuerdo")
            log.info(f"Opciones Acuerdo ({len(acuerdo_opts)}): {acuerdo_opts}")

            # Seleccionar el Acuerdo actual (o el primero disponible)
            catalogo_opts = []
            categoria_opts = []
            acuerdo_actual = self.acuerdo_var.get().strip()
            if acuerdo_opts:
                try:
                    sel_a = esperar_opciones_select(driver, "ajaxAcuerdo", WAIT_LARGO)
                    texto_a = acuerdo_actual if acuerdo_actual else acuerdo_opts[0]
                    seleccionar_por_texto_parcial(sel_a, texto_a)
                    time.sleep(2)
                    catalogo_opts = leer_opciones_select(driver, "ajaxCatalogo")
                    log.info(f"Opciones Catálogo ({len(catalogo_opts)}): {catalogo_opts}")
                except Exception as e:
                    log.warning(f"No se pudo cargar catálogos: {e}")

            # --- Catálogo → Categoría ---
            catalogo_actual = self.catalogo_var.get().strip()
            if catalogo_opts:
                try:
                    sel_c = esperar_opciones_select(driver, "ajaxCatalogo", WAIT_LARGO)
                    texto_c = catalogo_actual if catalogo_actual else catalogo_opts[0]
                    seleccionar_por_texto_parcial(sel_c, texto_c)
                    time.sleep(2)
                    categoria_opts = leer_opciones_select(driver, "ajaxCategoria")
                    log.info(f"Opciones Categoría ({len(categoria_opts)}): {categoria_opts}")
                except Exception as e:
                    log.warning(f"No se pudo cargar categorías: {e}")

            self.root.after(0, lambda: self._actualizar_combos(acuerdo_opts, catalogo_opts, categoria_opts))

        except Exception as e:
            err = str(e)
            log.error(f"Error cargando opciones del portal: {e}", exc_info=True)
            self.root.after(0, lambda: messagebox.showerror(
                "Error", f"No se pudieron cargar las opciones del portal:\n{err}"
            ))
            self.root.after(0, lambda: self.estado_var.set("Error al cargar opciones"))
        finally:
            if driver:
                try:
                    driver.quit()
                except Exception:
                    pass
            self.root.after(0, lambda: self.btn_cargar_opts.configure(state="normal"))
            self.root.after(0, lambda: self.btn_iniciar.configure(state="normal"))

    def _actualizar_combos(self, acuerdos, catalogos, categorias):
        self.combo_acuerdo["values"] = acuerdos
        self.combo_catalogo["values"] = catalogos
        self.combo_categoria["values"] = categorias

        # Asignar primer valor si el campo estaba vacío
        if acuerdos and not self.acuerdo_var.get():
            self.acuerdo_var.set(acuerdos[0])
        if catalogos and not self.catalogo_var.get():
            self.catalogo_var.set(catalogos[0])
        if categorias and not self.categoria_var.get():
            self.categoria_var.set(categorias[0])

        self._set_banner(
            f"✅  Filtros cargados — {len(acuerdos)} acuerdos, {len(catalogos)} catálogos, {len(categorias)} categorías",
            "#E8F5E9", "#2E7D32",
        )
        messagebox.showinfo(
            "Filtros cargados",
            f"Se cargaron desde el portal:\n"
            f"  • {len(acuerdos)} Acuerdo(s) Marco\n"
            f"  • {len(catalogos)} Catálogo(s)\n"
            f"  • {len(categorias)} Categoría(s)\n\n"
            f"Ahora selecciona los valores correctos en los desplegables del Paso 2.",
        )

    # ------------------------------------------------------------------
    # Pausa / Detección / Aprendizaje
    # ------------------------------------------------------------------
    def _pausar_reanudar(self):
        if not self._pausado:
            if PAUSA_EVENTO:
                PAUSA_EVENTO.clear()
            self._pausado = True
            self.btn_pausar.configure(text="▶  Reanudar")
            self._set_banner("⏸  En pausa — haz clic en 'Reanudar' para continuar", self.C_LOGIN_BG, "#7B5800")
            log.info("⏸ Ejecución pausada por el usuario.")
        else:
            if PAUSA_EVENTO:
                PAUSA_EVENTO.set()
            self._pausado = False
            self.btn_pausar.configure(text="⏸  Pausar")
            self._set_banner("▶  Reanudando proceso...", "#E8F0FE")
            log.info("▶ Ejecución reanudada por el usuario.")

    def _detener(self):
        if not messagebox.askyesno(
            "Detener proceso",
            "¿Seguro que quieres detener la automatización?\n\n"
            "Se generará el reporte Excel con los resultados hasta el momento.",
        ):
            return
        if DETENER_EVENTO:
            DETENER_EVENTO.set()
        if PAUSA_EVENTO:
            PAUSA_EVENTO.set()
        self._pausado = False
        self.btn_pausar.configure(text="⏸  Pausar")
        self._set_banner("⏹  Deteniendo — se terminará el producto actual y se generará el reporte",
                         "#FFF3E0", "#E65100")
        log.info("⏹ Detener solicitado por el usuario.")

    def _ver_aprendizaje(self):
        arch = Path(__file__).parent / "aprendizaje.json"
        if not arch.exists():
            messagebox.showinfo(
                "Estadísticas de errores",
                "Aún no hay datos registrados.\n\n"
                "El bot guarda estadísticas de cada error que encuentra.\n"
                "Después de la primera ejecución aparecerán aquí.",
            )
            return
        try:
            data = json.loads(arch.read_text(encoding="utf-8"))
            acum = data.get("acumulado", {})
            sesion = data.get("ultima_sesion", "Desconocida")
            if not acum:
                messagebox.showinfo("Estadísticas", "Aún no se han registrado errores.")
                return
            lineas = [
                f"Última sesión: {sesion}\n",
                "Errores encontrados (histórico):",
            ]
            for tipo, cnt in sorted(acum.items(), key=lambda x: -x[1]):
                estado = " ✔ ajuste activo" if cnt >= AnalizadorFallos.UMBRAL else ""
                lineas.append(f"  • {tipo}: {cnt} vez/veces{estado}")
            lineas += [
                "",
                f"Cuando un error se repite {AnalizadorFallos.UMBRAL}+ veces, el bot ajusta",
                "automáticamente su comportamiento para evitarlo.",
                "",
                "Para reiniciar las estadísticas: elimina el archivo 'aprendizaje.json'.",
            ]
            messagebox.showinfo("Estadísticas de errores", "\n".join(lineas))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo:\n{e}")

    def _iniciar(self):
        excel = Path(self.excel_var.get().strip())
        acuerdo = self.acuerdo_var.get().strip()
        catalogo = self.catalogo_var.get().strip()
        categoria = self.categoria_var.get().strip()
        pausa_txt = self.pausa_var.get().strip()

        if not excel.exists():
            messagebox.showerror(
                "Archivo no encontrado",
                f"No se encontró el archivo:\n{excel}\n\n"
                "Haz clic en '📂 Seleccionar archivo...' para elegir tu Excel."
            )
            return
        if not acuerdo or not catalogo or not categoria:
            messagebox.showerror(
                "Filtros incompletos",
                "Debes completar los tres filtros del Paso 2:\n"
                "  • Acuerdo Marco\n  • Catálogo\n  • Categoría\n\n"
                "Usa '🔄 Importar opciones del portal' si los desplegables están vacíos."
            )
            return
        try:
            pausa = int(pausa_txt)
            if pausa < 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Valor inválido", "La pausa debe ser un número entero mayor o igual a 0.")
            return

        # Resetear UI de resultado anterior
        self._mostrar_panel_resultado(False)
        self._mostrar_panel_ctrl(True)
        self._mostrar_panel_login(False)
        self._actualizar_progreso(0, 0, "Iniciando...")
        self.btn_iniciar.configure(state="disabled")
        self.btn_pausar.configure(text="⏸  Pausar")
        self._pausado = False
        self._set_banner("▶  Iniciando automatización — abriendo Chrome...", "#E8F0FE")

        self.worker = threading.Thread(
            target=self._worker_run,
            args=(excel, acuerdo, catalogo, categoria, pausa),
            daemon=True,
        )
        self.worker.start()

    def _worker_run(self, excel, acuerdo, catalogo, categoria, pausa):
        global MODO_GUI, EVENTO_LOGIN, GUI_NOTIFICAR_LOGIN, PAUSA_EVENTO, DETENER_EVENTO
        MODO_GUI = True
        EVENTO_LOGIN = self.login_event
        GUI_NOTIFICAR_LOGIN = self._notificar_login_ui
        PAUSA_EVENTO = threading.Event()
        PAUSA_EVENTO.set()
        DETENER_EVENTO = threading.Event()

        try:
            reporte = ejecutar_bot(
                excel_path=excel,
                acuerdo_texto=acuerdo,
                catalogo_texto=catalogo,
                categoria_texto=categoria,
                pausa_entre_productos=pausa,
            )
            self.reporte_generado = reporte
            total = len(RESULTADOS)
            exitos = sum(1 for r in RESULTADOS if r["Estado"] == "EXITO")
            fallos = total - exitos
            info = (
                f"{exitos} producto(s) actualizados correctamente"
                + (f"   ·   {fallos} con error(es)" if fallos else "")
                + f"\n\nReporte guardado en:\n{reporte}"
            )
            self.root.after(0, lambda: self._lbl_resultado_info.configure(text=info))
            self.root.after(0, lambda: self._mostrar_panel_ctrl(False))
            self.root.after(0, lambda: self._mostrar_panel_resultado(True))
            self.root.after(0, lambda: self._set_banner(
                f"✅  Proceso completado — {exitos}/{total} productos actualizados",
                "#E8F5E9", "#1B5E20",
            ))
            self.root.after(0, lambda: self._actualizar_progreso(total, total, "Completado"))
        except Exception as e:
            detalle = f"{e}\n\n{traceback.format_exc()}"
            log.error(f"Error fatal: {e}", exc_info=True)
            self.root.after(0, lambda: self._set_banner(
                "❌  Error en la ejecución — revisa la actividad detallada",
                "#FDECEA", self.C_PELIGRO,
            ))
            self.root.after(0, lambda: messagebox.showerror("Error inesperado", detalle))
        finally:
            self.root.after(0, lambda: self.btn_iniciar.configure(state="normal"))
            self._pausado = False


def iniciar_interfaz():
    root = tk.Tk()
    root.title("Peru Compras Bot")
    root.geometry("860x720")
    root.minsize(700, 560)
    root.resizable(True, True)

    root.withdraw()
    root.update_idletasks()
    root.deiconify()

    app = PeruComprasGUI(root)
    root.mainloop()


if __name__ == "__main__":
    if "--cli" in sys.argv:
        main_cli()
    else:
        iniciar_interfaz()
